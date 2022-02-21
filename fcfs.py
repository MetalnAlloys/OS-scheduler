#!/usr/bin/env python3

import time
import shutil
from enum import Enum
from typing import List, Dict
from openpyxl import load_workbook
from tabulate import tabulate


class State(Enum):
    IDLE = 0
    ACTIVE = 1
    EXPIRED = 2

# Make the output colorised
colors = {
    "green": "\033[1;32m",
    "off": "\033[0m",
    "red": "\033[0;31m",
    "blue": "\033[0;34m",
    "magenta": "\033[0;35m",
}
prefix = colors["green"] + "[+] " + colors["off"]
prefix_r = colors["red"] + "[-] " + colors["off"]
prefix_b = colors["blue"] + "[*] " + colors["off"]



def fcfs(tasks: List[Dict]):
    """ run tasks according to the FIFO scheme"""
    tu, nr_finished = 0, 0

    runqueue = []  # active tasks' list
    tasks = sorted(tasks, key=lambda i: i["arrival_time"])

    while nr_finished < len(tasks):
        if (nr_running := len(tasks)):
            for task in tasks:
                if task["arrival_time"] == tu:
                    runqueue.append(task)
                    nr_running += 1


        if len(runqueue) > 0 and runqueue[0]["burst_time"] == 0:
            print(prefix_r + "TU={}\tProcess {} finished execution".format(
                tu, runqueue[0]["pid"]))
            runqueue.pop(0)
            nr_finished += 1

            state = State.EXPIRED

        elif len(runqueue) > 0:
            runqueue[0]["burst_time"] -= 1
            time.sleep(0.05)

            for task in runqueue[1:]:
                task["waiting_time"] += 1

            state = State.ACTIVE
        else:
            state = State.IDLE

        display_execution(state, runqueue, tu)
        tu += 1



def display_execution(state, tasks, time):

    # case when CPU runs the Idle loop
    if state == State.IDLE or (state == State.EXPIRED and len(tasks) == 0):
        print(prefix_b + "TU={}\tCPU executing idle loop ...".format(time))

    # case when there is a process in queue                 
    elif state == State.ACTIVE or (state == State.EXPIRED and len(tasks) > 0):
        waiting = []
        for task in tasks[1:]:
            waiting.append("pid={}, wait={}".format(task["pid"], task["waiting_time"]))

        wait_que = ' '.join(waiting) if waiting else "<empty>"

        curr_pid = tasks[0]["pid"]
        rem_bt = tasks[0]["burst_time"]
        print(prefix +
              'TU={}\tPID={} executing - instructions left={}\tQueue: {}'.
              format(time, curr_pid, rem_bt, wait_que))


def display_tasks(tasks):
    headers = [
        "ID",
        "Arrival Time",
        "Burst Time",
        "Priority",
        "Waiting Time",
        "Turnaround Time",
    ]
    tasks_mat = []

    for task in tasks:
        tasks_mat.append(
            [
                task["pid"],
                f"{task['arrival_time']}",
                f"{task['burst_time']}",
                task["priority"],
                f"{task['waiting_time']}",
                f"{task['turnaround_time']}",
            ]
        )
    print(
        "\n"
        + tabulate(tasks_mat, headers=headers, tablefmt="fancy_grid", floatfmt=".3f")
    )
    


def get_task_list() -> List[Dict]:
    tasks = []

    wb = load_workbook('cpu-scheduling.xlsx') # assumes file is in cwd
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tasks.append({
            "pid": row[0],
            "arrival_time": row[1],
            "burst_time": row[2],
            "priority": row[3],
            "waiting_time": 0,
        })

    return tasks



if __name__ == "__main__":

    print("\n\n",u"\u001b[1m\u001b[4m\u001b[7m\t{}\t\u001b[0m\n" .format(

        "FIRST COME FIRST SERVE ALGORITHM"))

    term = shutil.get_terminal_size()
                                        
    print(prefix + "Loading tasks..\n")
    print('-' * term.columns)


    TASKS = get_task_list()


    fcfs(TASKS)
    display_tasks(TASKS)




